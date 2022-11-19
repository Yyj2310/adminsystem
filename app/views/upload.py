"""
文件上传
"""
from django.shortcuts import HttpResponse, render, redirect


def upload_list(request):
    """文件上传"""
    if request.method == 'GET':
        return render(request, 'upload_list.html')

    # print(request.POST)  # 上传的数据
    # print(request.FILES)  # 上传的文件
    file_object = request.FILES.get('avatar')
    # file_object.name  # 文件的名称

    # 获取文件：（文件会一块一块上传，需要一块一块得获取）
    f = open('a1.png', mode='wb')
    for chunk in file_object.chunks():
        f.write(chunk)
    f.close()
    return HttpResponse('...')


from openpyxl import load_workbook


def upload_multi(request):
    """批量提交"""
    # 获取上传的文件对象
    file_object = request.FILES.get('excel')
    # 把文件对象送给load_workbook→把文件直接打开，并读取内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    cell = sheet.cell(1, 2)
    print(cell.value)

    return HttpResponse('上传文件')
